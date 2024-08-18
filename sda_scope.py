# sda_scope.py

from PySide6.QtWidgets import QWidget, QVBoxLayout, QLabel, QSpacerItem, QSizePolicy, QLineEdit, QHBoxLayout, QCheckBox, QPushButton
from PySide6.QtCore import Qt
import os
import openpyxl

class SDAScope(QWidget):
    def __init__(self, app, sda_type):
        super().__init__()
        self.app = app
        self.sda_type = sda_type
        self.initUI()

    def initUI(self):
        main_layout = QVBoxLayout(self)

        # Top Layout für die Buttons
        top_layout = QVBoxLayout()
        top_layout.setAlignment(Qt.AlignLeft | Qt.AlignTop)

        # Beenden Button hinzufügen
        self.app.add_exit_button(top_layout)

        # Zurück Button hinzufügen
        back_button = self.app.create_button('Zurück')
        back_button.clicked.connect(lambda: self.app.show_page(3))  # Zurück zur IdentApproach-Seite
        top_layout.addWidget(back_button)

        # Center Layout für den Hauptinhalt
        center0_layout = QVBoxLayout()
        center0_layout.setAlignment(Qt.AlignCenter)

        # Label hinzufügen
        if self.sda_type == 'extended':
            label = QLabel(f"Select your Scope for SDA Extended")
        if self.sda_type == 'basic':
            label = QLabel(f"Select your Scope for SDA Fundamental")

        label.setStyleSheet('font-size: 50px; color: black;')

        # Spacer Item, um das Label höher zu positionieren
        #spacer = QSpacerItem(20, 100, QSizePolicy.Minimum, QSizePolicy.Expanding)
        #center_layout.addItem(spacer)

        center0_layout.addWidget(label, alignment=Qt.AlignCenter)

        # Spacer Item, um das Label weiter oben zu halten
        #spacer = QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Expanding)
        #center_layout.addItem(spacer)

        # Center Layout für den Hauptinhalt
        center_layout = QVBoxLayout()
        center_layout.setAlignment(Qt.AlignCenter)

        # Eingabefelder für Start year und End year
        input_layout = QVBoxLayout()
        input_layout.setAlignment(Qt.AlignCenter)

        self.input0 = QLineEdit(self)
        self.input0.setPlaceholderText("Start year")
        self.input0.setFixedWidth(200)
        self.input0.setStyleSheet("""
            font-size: 25px; 
            color: black; 
            background-color: #C63D2E; 
        """)
        input_layout.addWidget(self.input0)

        self.inputend = QLineEdit(self)
        self.inputend.setPlaceholderText("End year")
        self.inputend.setFixedWidth(200)
        self.inputend.setStyleSheet('font-size: 25px; color: black; background-color: #C63D2E;')
        input_layout.addWidget(self.inputend)

        # Eingabefelder zum Center Layout hinzufügen
        center_layout.addLayout(input_layout)

        # Checkbox-Listen für array1 und array2
        array1 = ["TBE", "CBE", "PBE", "BE"]

        current_path = os.getcwd()
        path1 = os.path.join(current_path, f'SDA_specs')
        excel_file_path = os.path.join(path1, f'iot_agg_region_names.xlsx')
        # Excel-Datei laden
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active
        erste_zeile = [cell.value for cell in ws[1]]
        # Ausgabe der Liste
        print(erste_zeile)
        #array2 = ["DE", "BR", "HDI-low", "HDI-med", "HDI-high"]
        array2 = erste_zeile

        excel_file_path = os.path.join(path1, f'iot_agg_sector_names.xlsx')
        # Excel-Datei laden
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active
        erste_zeile1 = [cell.value for cell in ws[1]]
        # Ausgabe der Liste
        print(erste_zeile1)
        #array2 = ["DE", "BR", "HDI-low", "HDI-med", "HDI-high"]
        array3 = erste_zeile1



        header_label1 = QLabel("Selection Demand Scheme")
        header_label1.setStyleSheet("font-size: 25px; color: black;")  # Stil der Überschrift anpassen

        # Checkbox-Listen für array1 und array2
        self.checkbox_layout1 = QVBoxLayout()
        self.checkbox_layout1.setAlignment(Qt.AlignCenter)
        self.checkbox_layout1.addWidget(header_label1, alignment=Qt.AlignCenter)
        self.checkbox_group1 = QWidget()
        self.checkbox_group1.setLayout(self.checkbox_layout1)
        #self.checkbox_group1.setStyleSheet('border: 1px solid black; padding: 5px;''QCheckBox::indicator { border: 2px solid black; }')
        self.checkbox_group1.setStyleSheet("""
        QCheckBox {
                font-size: 25px;  /* Schriftgröße der Beschreibung ändern */
            }
            QCheckBox::indicator {
                width: 25px;
                height: 25px;
            }
            QCheckBox::indicator:unchecked {
                border: 2px solid #000000;  /* Umrandung wenn nicht angekreuzt */
                background-color: #DADADA;  /* Hintergrundfarbe wenn nicht angekreuzt */
            }
            QCheckBox::indicator:checked {
                border: 2px solid #000000;  /* Umrandung wenn angekreuzt */
                background-color: #C63D2E;  /* Hintergrundfarbe wenn angekreuzt */
            }
        """)

        header_label2 = QLabel("Selection Regional Scheme")
        header_label2.setStyleSheet("font-size: 25px; color: black;")  # Stil der Überschrift anpassen

        self.checkbox_layout2 = QVBoxLayout()
        self.checkbox_layout2.setAlignment(Qt.AlignCenter)
        self.checkbox_layout2.addWidget(header_label2, alignment=Qt.AlignCenter)
        self.checkbox_group2 = QWidget()
        self.checkbox_group2.setLayout(self.checkbox_layout2)
        self.checkbox_group2.setStyleSheet("""
                QCheckBox {
                font-size: 25px;  /* Schriftgröße der Beschreibung ändern */
            }
            QCheckBox::indicator {
                width: 25px;
                height: 25px;
            }
            QCheckBox::indicator:unchecked {
                border: 2px solid #000000;  /* Umrandung wenn nicht angekreuzt */
                background-color: #DADADA;  /* Hintergrundfarbe wenn nicht angekreuzt */
            }
            QCheckBox::indicator:checked {
                border: 2px solid #000000;  /* Umrandung wenn angekreuzt */
                background-color: #C63D2E;  /* Hintergrundfarbe wenn angekreuzt */
            }
        """)

        header_label3 = QLabel("Selection Sectoral Aggregation")
        header_label3.setStyleSheet("font-size: 25px; color: black;")  # Stil der Überschrift anpassen

        # Checkbox-Listen für array1 und array2
        self.checkbox_layout3 = QVBoxLayout()
        self.checkbox_layout3.setAlignment(Qt.AlignCenter)
        self.checkbox_layout3.addWidget(header_label3, alignment=Qt.AlignCenter)
        self.checkbox_group3 = QWidget()
        self.checkbox_group3.setLayout(self.checkbox_layout3)
        # self.checkbox_group1.setStyleSheet('border: 1px solid black; padding: 5px;''QCheckBox::indicator { border: 2px solid black; }')
        self.checkbox_group3.setStyleSheet("""
         QCheckBox {
                 font-size: 25px;  /* Schriftgröße der Beschreibung ändern */
             }
             QCheckBox::indicator {
                 width: 25px;
                 height: 25px;
             }
             QCheckBox::indicator:unchecked {
                 border: 2px solid #000000;  /* Umrandung wenn nicht angekreuzt */
                 background-color: #DADADA;  /* Hintergrundfarbe wenn nicht angekreuzt */
             }
             QCheckBox::indicator:checked {
                 border: 2px solid #000000;  /* Umrandung wenn angekreuzt */
                 background-color: #C63D2E;  /* Hintergrundfarbe wenn angekreuzt */
             }
         """)


        self.checkboxes1 = []
        self.checkboxes2 = []
        self.checkboxes3 = []

        for item in array1:
            checkbox = QCheckBox(item)
            checkbox.setStyleSheet('font-size: 18px; color: black;')
            self.checkbox_layout1.addWidget(checkbox)
            self.checkboxes1.append(checkbox)

        for item in array2:
            checkbox = QCheckBox(item)
            checkbox.setStyleSheet('font-size: 18px; color: black;')
            self.checkbox_layout2.addWidget(checkbox)
            self.checkboxes2.append(checkbox)

        for item in array3:
            checkbox = QCheckBox(item)
            checkbox.setStyleSheet('font-size: 18px; color: black;')
            self.checkbox_layout3.addWidget(checkbox)
            self.checkboxes3.append(checkbox)

        # Hauptlayout für Checkbox-Listen
        checkbox_main_layout = QHBoxLayout()
        checkbox_main_layout.addWidget(self.checkbox_group1)
        checkbox_main_layout.addWidget(self.checkbox_group2)
        checkbox_main_layout.addWidget(self.checkbox_group3)

        # Checkbox-Listen zum Center Layout hinzufügen
        center_layout.addLayout(checkbox_main_layout)


        # Continue Button hinzufügen
        continue_button = QPushButton("Continue")
        continue_button.setStyleSheet('font-size: 25px; color: black; background-color: #C63D2E;')
        continue_button.setFixedSize(150, 50)  # Setze die feste Größe des Buttons (Breite, Höhe)
        continue_button.clicked.connect(self.on_continue)
        center_layout.addWidget(continue_button, alignment=Qt.AlignCenter)

        # Layouts zur Hauptlayout hinzufügen
        main_layout.addLayout(top_layout)
        main_layout.addStretch(1)
        main_layout.addLayout(center0_layout)
        main_layout.addStretch(1)
        main_layout.addLayout(center_layout)
        main_layout.addStretch(1)

        self.setLayout(main_layout)

    def on_continue(self):
        selected_demand = [cb.text() for cb in self.checkboxes1 if cb.isChecked()]
        selected_region = [cb.text() for cb in self.checkboxes2 if cb.isChecked()]
        selected_sector = [cb.text() for cb in self.checkboxes3 if cb.isChecked()]
        start_year = self.input0.text()
        end_year = self.inputend.text()
        self.app.show_ref_approach_sda_page(self.sda_type, start_year, end_year, selected_demand, selected_region, selected_sector)
