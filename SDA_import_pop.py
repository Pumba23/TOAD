import openpyxl
import os

def import_excel(year, category):

    current_path = os.getcwd()
    path3 = os.path.join(current_path, f'SDA_specs\population.xlsx')

    workbook = openpyxl.load_workbook(path3, data_only=True)
    sheet = workbook.active

    cellWW = sheet.cell(row=year - 1994, column=1)

    # Index, nach dem gesucht werden soll (z. B. "WW" oder "DE")
    index_to_search = category  # Beispiel-Index

    # Durchsuche die erste Zeile nach dem Index
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value == index_to_search:
            # Wenn ein Match gefunden wird, gib die Zelle darunter aus (zum Beispiel in der Zeile "year - 1994 + 1")
            result_cell = sheet.cell(row=year - 1994 + 1, column=col)
            print(f"Der Wert unter '{index_to_search}' in der Zeile für das Jahr {year} ist: {result_cell.value}")
            break
    else:
        print(f"Der Index '{index_to_search}' wurde nicht in der ersten Zeile gefunden.")

    #cellWW = sheet.cell(row=year - 1994, column=1)
    #cellDE = sheet.cell(row=year - 1994, column=2)


    popvall1 = result_cell.value
    popvall2 = cellWW.value

    print(popvall1)
    print(popvall2)

    return popvall1, popvall2
