
from mario import parse_exiobase_3
import numpy as np
import os
import openpyxl

def parse_aggr(year, category, sector):
    print('pathyear:')
    print(year)

    current_path = os.getcwd()
    path2 = os.path.join(current_path, f'SDA_specs')
    excel_file_path = os.path.join(path2, f'sector_names.xlsx')
    # Excel-Datei laden
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active

    erste_zeile = [cell.value for cell in ws[1]] #load sector names

    # Ausgabe der Liste
    print(erste_zeile)

    # Liste nach den Zahlen sortieren, die aus jeder Zeichenfolge extrahiert werden
    #sorted_sectors = ['Agriculture', 'Mining', 'Food', 'Non-energy-intensive manufacturing', 'Energy-intensive manufacturing', 'Electricity and gas', 'Services', 'Transport', 'Waste and wastewater supply']
    sorted_sectors = erste_zeile


    current_path = os.getcwd()
    path1 = os.path.join(current_path, f'EXIO3_IXI\IOT_' + str(year) + '_ixi.zip')
    world = parse_exiobase_3(path=path1) #load IOT


    #world.aggregate(r'C:\Users\VolkerHome\PycharmProjects\Mario_calc\EXIO3_IXI\iot_agg2.xlsx', ignore_nan = True) #aggregation for 1 reg and 9 secs

    current_path = os.getcwd()
    path2 = os.path.join(current_path, f'SDA')
    world.aggregate(path2+'_specs\iot_agg_sec_' + sector + '.xlsx', ignore_nan = True) #aggregation for sectors

    current_path = os.getcwd()
    path3 = os.path.join(current_path, f'SDA_specs')
    excel_file_path = os.path.join(path3, r'iot_agg_reg_' + category +'.xlsx')

    # Excel-Datei laden
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb['Sector']

    # Neue Spalte einfügen (zum Beispiel an der Stelle 2)
    column_index = 1
    # Check if the column is already inserted
    if ws.cell(row=1, column=column_index).value is None:
        ws.insert_cols(column_index)

    # Optional: Werte in die neue Spalte einfügen, adding names of sectors into regional aggregation
    new_column_values = sorted_sectors
    for row_num, value in enumerate(new_column_values, start=2):
        ws.cell(row=row_num, column=column_index, value=value)

    # Änderungen speichern
    wb.save(excel_file_path)


    #world.to_excel(r'C:\Users\VolkerHome\PycharmProjects\Mario_calc\abgleich2\dok ' + str(year) + '.xlsx',flows=True,coefficients=False) #saving of every years data

    print(world)

    current_path = os.getcwd()
    path1 = os.path.join(current_path, f'SDA')
    world.aggregate(path1+'_specs\iot_agg_reg_' + category + '.xlsx', ignore_nan = True) #aggregation for 1 reg and other parameters, sattelite account, factor of productein etc.

    print(world)


    # Excel-Datei laden
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb['Sector']  # Zugriff auf den Sheet "Sector"

    # Spalte definieren, aus der die Einträge gelöscht werden sollen (zum Beispiel Spalte 2)
    column_index = 1

    # Einträge ab der zweiten Zeile löschen (erste Zeile bleibt unberührt), delete the before added sector names
    for row_num in range(2, ws.max_row + 1):
        ws.cell(row=row_num, column=column_index).value = None

    # Änderungen speichern
    wb.save(excel_file_path)
    


    sectors_raw = world.get_index('Sector')  # Sectors
    v_sectors = np.array([sectors_raw]).flatten()
    num_sectors = v_sectors.shape[0]


    # In ein NumPy-Array umwandeln
    v_sectors = np.array(sorted_sectors)
    num_sectors = v_sectors.shape[0]

    regions_raw = world.get_index('Region')  # Sectors
    v_regions = np.array([regions_raw]).flatten()
    num_regions = v_regions.shape[0]

    return num_regions, num_sectors, world, v_sectors, v_regions
