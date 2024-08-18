import openpyxl
import os

def import_specs(sector):
    # Excel-Datei laden
    #print(sector)
    #print('----------------sector')
    current_path = os.getcwd()
    excel_file_path = os.path.join(current_path, r'IDA_data\IDA_' + sector + '_specs.xlsx')
    #print(excel_file_path)
    #print('----------------path')

    # Excel-Datei laden
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active

    # Variablen initialisieren
    start = end = legend_name = data_series = effects = None
    spaltennamen = []
    farben = []

    # Daten aus den ersten 5 Zeilen einlesen und Variablen zuweisen
    for row in range(1, 6):  # Erste 5 Zeilen
        key = ws[row][0].value
        value = ws[row][1].value
        if key == 'startyear':
            start = value
        elif key == 'endyear':
            end = value
        elif key == 'legendname':
            legend_name = value
        elif key == 'amount data series':
            data_series = value
        elif key == 'effects':
            effects = value

    # Anzahl der Spalten in Zeile 5 (für die letzten beiden Zeilen)
    num_cols_last_two_rows = len(ws[5])

    # Daten aus den letzten 2 Zeilen einlesen
    spaltennamen = [cell.value for cell in ws[6][1:num_cols_last_two_rows] if cell.value is not None]
    farben = [cell.value for cell in ws[7][1:num_cols_last_two_rows] if cell.value is not None]


    return start, end, legend_name, data_series, effects, spaltennamen, farben